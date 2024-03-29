import os
import sys
import time
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from db import crud_administration
from db import crud_form
from db import crud_question
from db import crud_user
from db import crud_data
from models.question import QuestionType
from models.answer import Answer
from models.user import UserRole
from db.connection import Base, SessionLocal, engine
from db.truncator import truncate
from source.geoconfig import GeoLevels
from source.corrections import Corrections

start_time = time.process_time()
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
Base.metadata.create_all(bind=engine)
session = SessionLocal()
source_path = os.environ["INSTANCE_NAME"]
SANDBOX_DATA_SOURCE = os.environ.get("SANDBOX_DATA_SOURCE")
if SANDBOX_DATA_SOURCE:
    source_path = SANDBOX_DATA_SOURCE
class_path = source_path.replace("-", "_")
source_file = "./source/baseline.xlsx"
all_sheets = load_workbook(source_file, read_only=True).sheetnames
sheets = list(filter(lambda x: source_path in x, all_sheets))
administration_level = [g["alias"] for g in GeoLevels[class_path].value]
lat_long = ["latitude", "longitude"]

corrections = {}

try:
    corrections = Corrections[class_path].value
except ValueError:
    print("No Data Corrections")

if len(sys.argv) < 2:
    print("You should provide admin address")
    sys.exit()

user = crud_user.get_user_by_email(session=session, email=sys.argv[1])
if not user:
    print("{} user is not available")
    sys.exit()

if user.role != UserRole.admin:
    print("{} user is not admin")
    sys.exit()


def get_location(sheet, x):
    parent_name = x[administration_level[0].lower()].strip()
    child_name = x[administration_level[1].lower()].strip()
    if parent_name in list(corrections):
        parent_name = corrections[parent_name]
    if child_name in list(corrections):
        child_name = corrections[child_name]
    parent = crud_administration.get_administration_by_name(
        session=session, name=parent_name
    )
    message = ""
    if not parent:
        message = "{}, {}, {} Not Found".format(sheet, x["ix"], parent_name)
        for pname in parent_name.split(" "):
            parent = crud_administration.get_administration_by_keyword(
                session=session, name=pname
            )
            if parent and not message:
                message = "{}, {},{} Not Found, perhaps {} ?".format(
                    sheet,
                    x["ix"],
                    parent_name,
                    " / ".join([p.name for p in parent]),
                )
        # raise ValueError(message)
        # print(message)
        return None
    children = crud_administration.get_administration_by_name(
        session=session, name=child_name, parent=parent.id
    )
    if children:
        return children.id
    message = ""
    for cname in child_name.split(" "):
        children = crud_administration.get_administration_by_keyword(
            session=session, name=cname, parent=parent.id
        )
        if children and not len(message):
            message = "{}, {}, {} Not Found, perhaps {} ?".format(
                sheet,
                x["ix"],
                child_name,
                " / ".join([c.name for c in children]),
            )
    if not len(message):
        message = "{}, {}, {} Not Found".format(sheet, x["ix"], child_name)
    # raise ValueError(message)
    # print(message)
    return None


def record(answers, form, user):
    administration = None
    geo = None
    answerlist = []
    names = []
    for a in answers:
        # Only get all non np.NaN value
        if answers[a] == answers[a]:
            valid = True
            qname = a.replace("_", " ").lower().strip()
            if "|" in qname:
                # to handle flow data
                qname = qname.split("|")[1]
            q = crud_question.get_question_by_name(
                session=session, name=qname, form=form.id
            )
            if not q:
                print(a)
            answer = Answer(question=q.id, created_by=user.id, created=datetime.now())
            if q.type == QuestionType.administration:
                administration = answers[a]
                answer.value = answers[a]
                if q.meta:
                    adm_name = crud_administration.get_administration_by_id(
                        session, id=administration
                    )
                    names.append(adm_name.name)
            if q.type == QuestionType.geo:
                if answers[a]:
                    geo = answers[a]
                    answer.text = ("{}|{}").format(answers[a][0], answers[a][1])
                else:
                    valid = False
            if q.type == QuestionType.text:
                answer.text = answers[a]
                if q.meta:
                    names.append(answers[a])
            if q.type == QuestionType.date:
                if answers[a]:
                    answer.text = answers[a]
                else:
                    valid = False
            if q.type == QuestionType.number:
                # changes: 0437ade9f8c1aa8ae6e1f23667d7975b958fd5a9
                try:
                    float(answers[a])
                    valid = True
                except ValueError:
                    valid = False
                if valid:
                    answer.value = answers[a]
                    if q.meta:
                        names.append(str(answers[a]))
            if q.type == QuestionType.option:
                answer.options = [answers[a]]
            if q.type == QuestionType.multiple_option:
                answer.options = answers[a]
            if valid:
                answerlist.append(answer)
    name = " - ".join([str(n) for n in names])
    data = crud_data.add_data(
        session=session,
        form=form.id,
        name=name,
        geo=geo,
        administration=administration,
        created_by=user.id,
        answers=answerlist,
    )
    return data


for table in ["data", "answer", "jmp_history", "history"]:
    action = truncate(session=session, table=table)
    print(action)

for sheet in sheets:
    data = pd.read_excel(source_file, sheet)
    data.drop(data.filter(regex="Unnamed"), axis=1, inplace=True)
    rename_columns = {"Latitude": "latitude", "Longitude": "longitude"}
    for adm in administration_level:
        rename_columns.update({adm: adm.lower()})
    data = data.rename(columns=rename_columns)
    data.dropna(subset=[a.lower() for a in administration_level], inplace=True)
    data["ix"] = data.index + 1
    data["location"] = data.apply(lambda x: get_location(sheet, x), axis=1)
    data = data.drop(columns=["ix"], axis=1)
    data = data.dropna(subset=["location"])
    if data.shape[0]:
        for adm in administration_level:
            data.drop(data.filter(regex=adm.lower()), axis=1, inplace=True)
        geo = False
        for g in lat_long:
            geo = True
            if g not in list(data):
                geo = False
        if geo:
            data["geolocation"] = data.apply(
                lambda x: [x["latitude"], x["longitude"]]
                if x["latitude"] == x["latitude"] and x["longitude"] == x["longitude"]
                else None,
                axis=1,
            )
            data = data.drop(columns=lat_long, axis=1)
        form_name = sheet.replace(source_path, "").strip()
        form = crud_form.search_form_by_name(session=session, name=form_name)
        data = data.to_dict("records")
        for d in data:
            rec = record(d, form, user)

elapsed_time = time.process_time() - start_time
elapsed_time = str(timedelta(seconds=elapsed_time)).split(".")[0]
print(f"\n-- DONE IN {elapsed_time}\n")
